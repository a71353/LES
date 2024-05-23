from collections import defaultdict
from django.shortcuts import render
from configuracao.models import Diaaberto,Menu, Prato
from questionariosPublicados.models import *
# Create your views here.
import json
import csv
from django.http import HttpResponse
from configuracao.models import Diaaberto
from questionariosPublicados.models import Pergunta, Resposta_Individual
from django.contrib.auth import get_user
from django.shortcuts import render, redirect, get_object_or_404

def estatisticas_almocos(request):
    user = get_user(request)
    if user.groups.filter(name="Administrador").exists():
        dias_abertos = Diaaberto.objects.all()
        nome_questionario = None
        data_publicacao = None
        data_expiracao = None
        respostas_json_escolhas = []
        respostas_almocos_escolhas_multiplas = []
        respostas_json = []
        perguntas_almocos = []
        respostas= ""
        participantes= 0
        ano_diaaberto = None
        if request.method == 'POST':
            diaabertoID = request.POST.get('diaaberto')
            print("Gustavo: ",diaabertoID)
            if diaabertoID != "":
                
                diaaberto = Diaaberto.objects.get(id=diaabertoID)
                ano_diaaberto = diaaberto.ano #ano do dia aberto
                if diaaberto.questionario == None: #n tÊm questionarios associados
                    respostas= "No"
                    print(respostas)
                    return redirect('utilizadores:mensagem', 9002)
                else:
                    nome_questionario = diaaberto.questionario.nome
                    questionarioId = diaaberto.questionario.id
                    data_publicacao = diaaberto.questionario.data_publicada
                    data_expiracao = diaaberto.questionario.data_fim_publicacao
                    perguntasquestionario = Pergunta.objects.filter(questionario=questionarioId).all()
                    
                    temaAlmoco = False
                    for perg in perguntasquestionario:
                        if perg.tema.tema == "Almoço":
                            perguntas_almocos.append(perg)
                            temaAlmoco = True
                            
                    if temaAlmoco == True:
                        get_respostas = Resposta.objects.filter(questionario=questionarioId).all()
                        if get_respostas.count() == 0: # n há respostas
                            respostas= "False"
                            return redirect('utilizadores:mensagem', 9003)
                        else: #Calcular o numero de respostas
                            respostas= "True"
                            participantes = get_respostas.count()                
                            respostas_almocos = []
                            respostas_almocos_escolhas_multiplas = []
                            id = 0
                            for pergunta in perguntas_almocos:
                                resposta_alm = {}
                                resposta_alm['id']  = id
                                resposta_alm['pergunta'] = pergunta.texto
                                resposta_alm['tipo_pergunta'] = pergunta.tipo_resposta
                                resposta_alm['respostas'] = []
                                if pergunta.tipo_resposta == 'multipla_escolha':
                                    opcoes = pergunta.opcoes.all()
                                    
                                    for opcao in opcoes:
                                        # Contar quantas vezes esta opção foi selecionada
                                        count_opcao = Resposta_Individual.objects.filter(
                                            resposta__in=get_respostas,
                                            pergunta=pergunta,
                                            resposta_texto=str(opcao.id)  
                                        ).count()

                                        percentagem_opcao = (count_opcao / participantes) * 100
                                        

                                        resposta_alm['respostas'].append({
                                            'opcao': opcao.texto_opcao,
                                            'count': percentagem_opcao
                                        })
                                    respostas_almocos_escolhas_multiplas.append(resposta_alm)
                                    
                                    id = id + 1
                        
                                respostas_almocos.append(resposta_alm)

                                

                            respostas_json = json.dumps(respostas_almocos) 
                            respostas_json_escolhas = json.dumps(respostas_almocos_escolhas_multiplas) 
                                    
                            
        
            return render(request, 'almocos/estatisticas_almocos.html',
                    context={'diasAbertos':dias_abertos,
                            'participantes':participantes,'respostas':respostas,
                            'respostas_almocos':respostas_json,
                            'respostas_almocos_escolhas_multiplas': respostas_almocos_escolhas_multiplas
                            ,'respostas_escolhas_multiplas_json':respostas_json_escolhas
                            ,'data_publicacao':data_publicacao
                            ,'data_expiracao':data_expiracao
                            ,'ano_diaaberto':ano_diaaberto
                            ,'diaabertoID':diaabertoID
                            ,'nome_questionario': nome_questionario});
        else:
            return render(request, 'almocos/estatisticas_almocos.html',
                    context={'diasAbertos':dias_abertos});
    else:
        return redirect('utilizadores:mensagem',9000)

def estatisticas_almocos_csv(request):
    user = get_user(request)
    if user.groups.filter(name="Administrador").exists():

        diaabertoID = request.GET.get('diaaberto')  # Pegue o ID do Dia Aberto do parâmetro da URL
        
        # Criar a resposta HTTP com tipo de conteúdo CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="estatisticas_almocos.csv"'
        
        writer = csv.writer(response)
        
        if diaabertoID:
            try:
                diaaberto = Diaaberto.objects.get(id=diaabertoID)
            except Diaaberto.DoesNotExist:
                return HttpResponse("Dia Aberto não encontrado.", content_type="text/plain")
            
            if diaaberto.questionario:
                questionarioId = diaaberto.questionario.id
                perguntas = Pergunta.objects.filter(questionario=questionarioId, tema__tema="Almoço", tipo_resposta='multipla_escolha').all()
                
                if perguntas:
                    headers = ['Pergunta', 'Opções', 'Total', 'Porcentagem']  # Cabeçalhos atualizados sem 'Tema'
                    writer.writerow(headers)

                    for pergunta in perguntas:
                        opcoes = pergunta.opcoes.all()
                        
                        total_respostas = Resposta_Individual.objects.filter(pergunta=pergunta).count()
                        
                        for index, opcao in enumerate(opcoes):
                            count_opcao = Resposta_Individual.objects.filter(pergunta=pergunta, resposta_texto=str(opcao.id)).count()
                            porcentagem = f"{(count_opcao / total_respostas * 100) if total_respostas > 0 else 0:.2f}%"
                            
                            if index == 0:
                                writer.writerow([pergunta.texto, opcao.texto_opcao, count_opcao, porcentagem])
                            else:
                                writer.writerow(['', opcao.texto_opcao, count_opcao, porcentagem])
                        
                        writer.writerow([])  # Linha vazia após cada pergunta para separação visual
                else:
                    writer.writerow(['Não há perguntas de múltipla escolha para o tema "Almoço" associadas a este questionário.'])
            else:
                return HttpResponse("Não há questionário associado ao Dia Aberto selecionado.", content_type="text/plain")
        else:
            return HttpResponse("ID do Dia Aberto não fornecido ou inválido.", content_type="text/plain")
        
        return response
    else:
        return redirect('utilizadores:mensagem',9001)
    
from django.http import HttpResponse
from openpyxl import Workbook
from django.shortcuts import redirect

def estatisticas_almocos_excel(request):
    user = get_user(request)
    if not user.groups.filter(name="Administrador").exists():
        return redirect('utilizadores:mensagem', 9001)

    diaabertoID = request.GET.get('diaaberto')
    if not diaabertoID:
        return HttpResponse("ID do Dia Aberto não fornecido ou inválido.", content_type="text/plain")

    try:
        diaaberto = Diaaberto.objects.get(id=diaabertoID)
    except Diaaberto.DoesNotExist:
        return HttpResponse("Dia Aberto não encontrado.", content_type="text/plain")

    if not diaaberto.questionario:
        return HttpResponse("Não há questionário associado ao Dia Aberto selecionado.", content_type="text/plain")

    questionarioId = diaaberto.questionario.id
    perguntas = Pergunta.objects.filter(questionario=questionarioId, tema__tema="Almoço", tipo_resposta='multipla_escolha').all()
    if not perguntas:
        return HttpResponse("Não há perguntas de múltipla escolha para o tema 'Almoço' associadas a este questionário.", content_type="text/plain")

    wb = Workbook()
    wb.remove(wb.active)  # remove a aba padrão criada automaticamente

    contador = 0
    for pergunta in perguntas:
        contador += 1
        if contador == 1:
            sheet_title = "Pergunta"
        else:
            sheet_title = f"Pergunta{contador}"
        ws = wb.create_sheet(title=sheet_title)
        headers = ['Pergunta', 'Opções', 'Total', 'Porcentagem']
        ws.append(headers)

        opcoes = pergunta.opcoes.all()
        total_respostas = Resposta_Individual.objects.filter(pergunta=pergunta).count()
        first_line = True

        for opcao in opcoes:
            count_opcao = Resposta_Individual.objects.filter(pergunta=pergunta, resposta_texto=str(opcao.id)).count()
            porcentagem = f"{(count_opcao / total_respostas * 100) if total_respostas > 0 else 0:.2f}%"
            if first_line:
                ws.append([pergunta.texto, opcao.texto_opcao, count_opcao, porcentagem])
                first_line = False
            else:
                ws.append(['', opcao.texto_opcao, count_opcao, porcentagem])

        # Ajuste automático da largura das colunas
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[ws.cell(row=1, column=col[0].column).column_letter].width = max_length

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="estatisticas_almocos.xlsx"'
    wb.save(response)

    return response

def estatistica_transporte(request):
    user = get_user(request)
    if not user.groups.filter(name="Administrador").exists():
        return redirect('utilizadores:mensagem', 9000)

    diaabertoid = request.GET.get('diaaberto')
    if not diaabertoid:
        diaabertoid = Diaaberto.objects.filter(ano__lte=datetime.now().year).order_by('-ano').first().id
    diaaberto = get_object_or_404(Diaaberto, id=diaabertoid)

    if not diaaberto.questionario_id:
        return redirect('utilizadores:mensagem1', 19)

    questionario = get_object_or_404(Questionario, id=diaaberto.questionario_id)
    tema_transporte = get_object_or_404(TemaQ, tema='Transporte')
    perguntas = Pergunta.objects.filter(questionario_id=questionario.id, tema=tema_transporte.id, tipo_resposta='multipla_escolha')

    resultados = []
    if perguntas.exists():
        for pergunta in perguntas:
            respostas_individuais = Resposta_Individual.objects.filter(pergunta=pergunta)
            opcoes = OpcaoP.objects.filter(pergunta=pergunta).in_bulk(field_name='id')
            contador_respostas = defaultdict(int)

            for resposta in respostas_individuais:
                id_opcao = resposta.resposta_texto
                texto_opcao = opcoes[int(id_opcao)].texto_opcao if int(id_opcao) in opcoes else 'Opção não encontrada'
                contador_respostas[texto_opcao] += 1

            respostas_agrupadas = [{'texto_opcao': texto_opcao, 'total': total} for texto_opcao, total in contador_respostas.items()]
            resultados.append({
                'pergunta': pergunta.texto,
                'respostas': respostas_agrupadas
            })

    return render(request, 'estatisticas/estatistica_transporte.html', {
        'diaaberto': diaaberto,
        'diasabertos': Diaaberto.objects.all(),
        'resultados': resultados
    })