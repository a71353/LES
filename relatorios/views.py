import csv
from itertools import count, groupby
import time
from django.urls import reverse
from django.utils import timezone
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import redirect, render
from django_tables2 import SingleTableMixin

from inscricoes.models import Escola, Inscricao, Inscricaoprato, Inscricaosessao, Inscricaotransporte
from inscricoes.utils import render_pdf
from relatorios.filters import RelatorioFilter
from utilizadores.models import Administrador
from .models import Relatorio
from .tables import RelatorioTable
from django_filters.views import FilterView
from configuracao.models import Diaaberto
from django.db.models import Sum
from .forms import RelatorioForm
from django.contrib import messages
from django.db import OperationalError, transaction
from django.contrib.auth.mixins import LoginRequiredMixin
import os
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from django.http import HttpResponse, FileResponse
import os
import io
from django.contrib.staticfiles.storage import staticfiles_storage
from urllib.parse import quote
from datetime import datetime
from collections import defaultdict
from django.http import HttpResponse
import csv
from django.db.models import Q
from django.http import JsonResponse

def handle_db_errors(view_func):
    def wrapper(request, *args, **kwargs):
        try:
            response = view_func(request, *args, **kwargs)
            return response
        except OperationalError as e:
            print(f"Database error encountered: {e}")
            return render(request, "db_error.html", status=503)
    return wrapper



BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RELATORIOS_DIR = os.path.join(BASE_DIR, 'relatorios_internos')

if not os.path.exists(RELATORIOS_DIR):
    os.makedirs(RELATORIOS_DIR)


class ListaRelatorios(SingleTableMixin, FilterView):
    model = Relatorio
    table_class = RelatorioTable
    template_name = 'relatorios/lista_relatorios.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        relatorios_path = os.path.join(BASE_DIR, 'relatorios_internos')

        relatorios = []
        for f in os.listdir(relatorios_path):
            full_path = os.path.join(relatorios_path, f)
            if os.path.isfile(full_path):
                # Assuming the filename format is "type_date_time.pdf"
                parts = f.split('_')
                report_type = parts[0]  # Type is the first part before the first underscore
                relatorios.append({
                    'name': f,
                    'type': report_type.capitalize(),  # Capitalize or format as needed
                    'url_safe_name': quote(f)
                })

        context['relatorios'] = relatorios
        context['has_relatorios'] = bool(relatorios)  # True if there are reports, False otherwise
        return context

@handle_db_errors
def create_report(request, dia_aberto, export_type, tema):
    relatorio_data = {
        'tema': tema,
        'dia_aberto': dia_aberto.id,
        'autor': request.user.id,
        'data_exportado': timezone.now().date(),
        'export_type': export_type
    }
    form = RelatorioForm(relatorio_data)
    if form.is_valid():
        form.save()
        messages.success(request, f"Relatório {export_type} criado com sucesso!")
    else:
        for error in form.errors.values():
            messages.error(request, error)

@handle_db_errors
def relatorio_transporte(request, ano=None, rep=None):
    """View que gera um PDF com os detalhes das inscrições para um dado ano do diaaberto"""

    dia_aberto = Diaaberto.objects.get(ano=ano)
    inscricoes = Inscricao.objects.filter(diaaberto=dia_aberto).select_related('escola')

    if not inscricoes.exists():
        return redirect('utilizadores:mensagem2', 20)

    if rep == 'yes':
        create_report(request, dia_aberto, 'PDF', 'Transporte')

    transportes_inscricoes = {}
    for inscricao in inscricoes:
        escola = inscricao.escola.nome
        dia_inscricao = inscricao.dia
        ano_inscricao = inscricao.ano  # Acessando o ano da inscrição
        turma_inscricao = inscricao.turma  # Acessando a turma da inscrição
        transportes = Inscricaotransporte.objects.filter(inscricao=inscricao)
        if not transportes.exists():
            return redirect('utilizadores:mensagem6', 1)
        if dia_inscricao not in transportes_inscricoes:
            transportes_inscricoes[dia_inscricao] = []

        for transporte_inscricao in transportes:
            transporte_info = {
                'inscricao_id': inscricao.id,
                'meio_transporte': inscricao.meio_transporte,
                'npassageiros': transporte_inscricao.npassageiros,
                'capacidade': transporte_inscricao.transporte.get_capacidade,
                'escola_nome': escola,
                'turma_inscricao': inscricao.turma,
                'ano_inscricao': ano_inscricao,  
                'turma_inscricao': turma_inscricao,  
                'hora_partida': transporte_inscricao.transporte.horaPartida,
                'hora_chegada': transporte_inscricao.transporte.horaChegada,
                'origem': transporte_inscricao.transporte.origem,
                'chegada': transporte_inscricao.transporte.chegada,
            }
            transportes_inscricoes[dia_inscricao].append(transporte_info)

    context = {
        'transportes_inscricoes': transportes_inscricoes,
        'ano': ano,
    }

    filename = f"transportes_dia_aberto_{ano}.pdf"
    if rep == 'yes':
        return render_pdf(request, "relatorios/relatorio_transporte.html", context, filename, save_file=True)
    else:
        return render_pdf(request, "relatorios/relatorio_transporte.html", context, filename)
    
@handle_db_errors
def relatorio_transporte_dia(request, ano=None, rep=None, day=None):
    """View que gera um PDF com os detalhes das inscrições para um dado ano e dia do diaaberto"""

    dia_aberto = Diaaberto.objects.get(ano=ano)
    # Filtra inscrições por dia e ano
    inscricoes = Inscricao.objects.filter(diaaberto=dia_aberto, dia=day).select_related('escola')

    if not inscricoes.exists():
        return redirect('utilizadores:mensagem6', 1)

    if rep == 'yes':
        create_report(request, dia_aberto, 'PDF', 'Transporte')

    transportes_inscricoes = {}
    for inscricao in inscricoes:
        escola = inscricao.escola.nome
        transportes = Inscricaotransporte.objects.filter(inscricao=inscricao)
        if not transportes.exists():
            return redirect('utilizadores:mensagem2', 20)
        if day not in transportes_inscricoes:
            transportes_inscricoes[day] = []

        for transporte_inscricao in transportes:
            transporte_info = {
                'inscricao_id': inscricao.id,
                'meio_transporte': inscricao.meio_transporte,
                'npassageiros': transporte_inscricao.npassageiros,
                'capacidade': transporte_inscricao.transporte.get_capacidade,
                'escola_nome': escola,
                'turma_inscricao': inscricao.turma,
                'ano_inscricao': inscricao.ano,  
                'hora_partida': transporte_inscricao.transporte.horaPartida,
                'hora_chegada': transporte_inscricao.transporte.horaChegada,
                'origem': transporte_inscricao.transporte.origem,
                'chegada': transporte_inscricao.transporte.chegada,
            }
            transportes_inscricoes[day].append(transporte_info)

    context = {
        'transportes_inscricoes': transportes_inscricoes,
        'ano': ano,
        'dia': day
    }

    filename = f"transportes_dia_aberto_{ano}_{day}.pdf"
    if rep == 'yes':
        return render_pdf(request, "relatorios/relatorio_transporte.html", context, filename, save_file=True)
    else:
        return render_pdf(request, "relatorios/relatorio_transporte.html", context, filename)

def relatorio_transporte_csv(request, ano=None, rep=None):
    try:
        dia_aberto = Diaaberto.objects.get(ano=ano)
        inscricoes = Inscricao.objects.filter(diaaberto_id=dia_aberto.id)

        if not inscricoes.exists():
            return redirect('utilizadores:mensagem2', 20)

        transportes_existentes = Inscricaotransporte.objects.filter(inscricao__in=inscricoes).exists()
        if not transportes_existentes:
            return redirect('utilizadores:mensagem6', 1)

        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';')
        # Header with the year of the event
        writer.writerow(['Relatório de Transportes do Dia Aberto', ano])
        writer.writerow(['Inscricao ID', 'Escola', 'Meio de Transporte', 'Origem', 'Chegada', 'Hora Partida', 'Hora Chegada', 'Capacidade', 'Numero de Passageiros'])

        inscricoes_por_dia = {}
        for inscricao in inscricoes:
            dia_inscricao = inscricao.dia  
            if dia_inscricao not in inscricoes_por_dia:
                inscricoes_por_dia[dia_inscricao] = []
            inscricoes_por_dia[dia_inscricao].append(inscricao)

        for dia, inscricoes_dia in inscricoes_por_dia.items():
            populate_csv_transportes(writer, dia, inscricoes_dia)

        if rep == 'yes':
            filepath = os.path.join(RELATORIOS_DIR, f"transportes_dia_aberto_{ano}.csv")
            with open(filepath, 'w', newline='', encoding='utf-8') as file:
                file.write(buffer.getvalue())

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="transportes_dia_aberto_{ano}.csv"'
        return response
    except Diaaberto.DoesNotExist:
        return redirect('utilizadores:mensagem2', 21)

def populate_csv_transportes(writer, dia, inscricoes_dia):
    writer.writerow([])
    writer.writerow([f'Dia {dia}'])
    writer.writerow(['Inscricao ID', 'Escola', 'Meio de Transporte', 'Origem', 'Chegada', 'Hora Partida', 'Hora Chegada', 'Capacidade', 'Numero de Passageiros'])
    for inscricao in inscricoes_dia:
        transportes = Inscricaotransporte.objects.filter(inscricao=inscricao)
        for transporte_inscricao in transportes:
            transporte_horario = transporte_inscricao.transporte
            writer.writerow([
                inscricao.id,
                inscricao.escola.nome,
                inscricao.meio_transporte,
                transporte_horario.origem,
                transporte_horario.chegada,
                transporte_horario.horaPartida,
                transporte_horario.horaChegada,
                transporte_horario.get_capacidade,
                transporte_inscricao.npassageiros
            ])
@handle_db_errors
def relatorio_transporte_csv_dia(request, ano=None, rep=None, day=None):
    """View que gera um CSV com os detalhes das inscrições de transporte para um dado ano e dia do diaaberto"""

    dia_aberto = Diaaberto.objects.get(ano=ano)
    # Filtra inscrições por ano e dia especificado
    inscricoes = Inscricao.objects.filter(diaaberto_id=dia_aberto.id, dia=day)

    if not inscricoes.exists():
        return redirect('utilizadores:mensagem2', 20)

    
    transportes_existentes = Inscricaotransporte.objects.filter(inscricao__in=inscricoes).exists()
    if not transportes_existentes:
        return redirect('utilizadores:mensagem6', 1)

    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=';')
    writer.writerow(['', 'Inscricao ID', 'Escola', 'Meio de Transporte', 'Origem', 'Chegada', 'Hora Partida', 'Hora Chegada', 'Capacidade', 'Numero de Passageiros'])

    if inscricoes:
        writer.writerow([f'Dia {day}'])  # Escreve uma linha para o dia
        for inscricao in inscricoes:
            transportes = Inscricaotransporte.objects.filter(inscricao=inscricao)
            for transporte_inscricao in transportes:
                transporte_horario = transporte_inscricao.transporte
                writer.writerow([
                    '',  
                    inscricao.id,
                    inscricao.escola.nome,
                    inscricao.meio_transporte,
                    transporte_horario.origem,
                    transporte_horario.chegada,
                    transporte_horario.horaPartida,
                    transporte_horario.horaChegada,
                    transporte_horario.get_capacidade,
                    transporte_inscricao.npassageiros
                ])

    if rep == 'yes':
        filepath = os.path.join(RELATORIOS_DIR, f"transportes_dia_aberto_{ano}_{day}.csv")
        with open(filepath, 'w', newline='', encoding='utf-8') as file:
            file.write(buffer.getvalue())

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="transportes_dia_aberto_{ano}_{day}.csv"'
    return response




@handle_db_errors
def render_pdf(request, template_src, context, output_filename, save_file=False):
    # Suponha que você queira gerar um URL absoluto para uma imagem estática
    logo_path = staticfiles_storage.url('img/ualg-logo.png')
    logo_path2 = staticfiles_storage.url('img/logo-large.png')
    context['logo_url'] = request.build_absolute_uri(logo_path)
    context['logo_url2'] = request.build_absolute_uri(logo_path2)

    html = render_to_string(template_src, context)
    result = io.BytesIO()
    pdf = pisa.CreatePDF(io.BytesIO(html.encode("UTF-8")), result)

    if pdf.err:
        return HttpResponse("Erro ao gerar o PDF.", status=500)
    
    if save_file:
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename_with_timestamp = f"{output_filename.split('.')[0]}_{timestamp}.pdf"
        file_path = os.path.join(RELATORIOS_DIR, filename_with_timestamp)
        with open(file_path, 'wb') as pdf_file:
            pdf_file.write(result.getvalue())

    return FileResponse(io.BytesIO(result.getvalue()), content_type='application/pdf', as_attachment=True, filename=output_filename)


@handle_db_errors
def relatorio_atividade(request, ano=None, rep=None):
    dia_aberto = Diaaberto.objects.get(ano=ano)
    inscricoes = Inscricao.objects.filter(diaaberto=dia_aberto)
    if not inscricoes.exists():
        return redirect('utilizadores:mensagem2', 20)

    sessoes_data = (Inscricaosessao.objects
                    .filter(inscricao__in=inscricoes)
                    .select_related('sessao__atividadeid', 'sessao__horarioid')
                    .values('sessao__atividadeid__nome', 'sessao__atividadeid__tipo', 
                            'sessao__horarioid__inicio', 'sessao__horarioid__fim', 'sessao__dia', 'sessao__atividadeid__participantesmaximo')
                    .annotate(nparticipantes_total=Sum('nparticipantes'))
                    .order_by('sessao__horarioid__inicio', 'sessao__dia'))

    if not sessoes_data.exists():
        return redirect('utilizadores:mensagem6', 1)
    # Grouping sessions by day
    grouped_sessoes_data = []
    for key, group in groupby(sessoes_data, key=lambda x: x['sessao__dia']):
        grouped_sessoes_data.append({
            'day': key,
            'sessions': list(group)
        })

    inscricoes_por_dia = defaultdict(list)
    for inscricao in inscricoes:
        inscricoes_por_dia[inscricao.dia].append(inscricao)

    context = {
        'ano': ano,
        'grouped_sessoes_data': grouped_sessoes_data,
    }

    filename = f"atividades_dia_aberto_{ano}.pdf"
    if rep == 'yes':
        # Salvar o PDF internamente e retorná-lo
        return render_pdf(request, "relatorios/relatorio_atividade.html", context, filename, save_file=True)
    else:
        # Apenas gerar o PDF para visualização, sem salvar
        return render_pdf(request, "relatorios/relatorio_atividade.html", context, filename)
    
@handle_db_errors
def relatorio_atividade_dia(request, ano=None, rep=None, day=None):
    try:
        dia_aberto = Diaaberto.objects.get(ano=ano)
        inscricoes = Inscricao.objects.filter(diaaberto=dia_aberto)
        if not inscricoes.exists():
            return redirect('utilizadores:mensagem2', 20)

        # Filtrando sessões por dia, se especificado
        if day:
            sessoes_data = (Inscricaosessao.objects
                            .filter(inscricao__in=inscricoes, sessao__dia=day)
                            .select_related('sessao__atividadeid', 'sessao__horarioid'))
        else:
            sessoes_data = (Inscricaosessao.objects
                            .filter(inscricao__in=inscricoes)
                            .select_related('sessao__atividadeid', 'sessao__horarioid'))

        sessoes_data = sessoes_data.values(
            'sessao__atividadeid__nome', 'sessao__atividadeid__tipo',
            'sessao__horarioid__inicio', 'sessao__horarioid__fim', 'sessao__dia', 'sessao__atividadeid__participantesmaximo'
        ).annotate(nparticipantes_total=Sum('nparticipantes')).order_by('sessao__horarioid__inicio', 'sessao__dia')

        if not sessoes_data.exists():
            return redirect('utilizadores:mensagem6', 1)

        # Agrupando sessões por dia
        grouped_sessoes_data = []
        for key, group in groupby(sessoes_data, key=lambda x: x['sessao__dia']):
            grouped_sessoes_data.append({
                'day': key,
                'sessions': list(group)
            })

        context = {
            'ano': ano,
            'grouped_sessoes_data': grouped_sessoes_data,
        }

        filename = f"atividades_dia_aberto_{ano}.pdf"
        if rep == 'yes':
            # Salvar o PDF internamente e retorná-lo
            return render_pdf(request, "relatorios/relatorio_atividade.html", context, filename, save_file=True)
        else:
            # Apenas gerar o PDF para visualização, sem salvar
            return render_pdf(request, "relatorios/relatorio_atividade.html", context, filename)

    except Diaaberto.DoesNotExist:
        return JsonResponse({'error': 'No open day found for the given year.'}, status=404)
    

@handle_db_errors
def relatorio_atividade_csv(request, ano=None, rep=None):
    """View que gera e opcionalmente salva um CSV com detalhes das sessões para um dado ano do Dia Aberto."""
    dia_aberto = Diaaberto.objects.get(ano=ano)
    inscricoes = Inscricao.objects.filter(diaaberto=dia_aberto)

    if not inscricoes.exists():
        return redirect('utilizadores:mensagem2', 20)

    sessoes = Inscricaosessao.objects.filter(inscricao__in=inscricoes)
    if not sessoes.exists():
        return redirect('utilizadores:mensagem6', 1)  # Assumindo que esta URL trata o caso sem sessões


    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=';')
    writer.writerow(['Atividade', 'Tipo', 'Data da Sessao', 'Hora de Inicio', 'Hora de Termino', 'Numero de Participantes', 'Numero Maximo de Participantes'])
    populate_csv(writer, inscricoes)

    # Se a opção 'rep' for 'yes', salvar o arquivo internamente
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")  # Gerar timestamp
    if rep == 'yes':
        filepath = os.path.join(RELATORIOS_DIR, f"atividades_dia_aberto_{ano}_{timestamp}.csv")
        with open(filepath, 'w', newline='', encoding='utf-8') as file:
            file.write(buffer.getvalue())

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="atividades_dia_aberto_{ano}_{timestamp}.csv"'
    return response
@handle_db_errors
def relatorio_atividade_csv_dia(request, ano=None, rep=None, day=None):
    """View que gera e opcionalmente salva um CSV com detalhes das sessões para um dado ano e dia do Dia Aberto."""
    try:
        dia_aberto = Diaaberto.objects.get(ano=ano)
        inscricoes = Inscricao.objects.filter(diaaberto=dia_aberto)

        if not inscricoes.exists():
            return redirect('utilizadores:mensagem2', 20)

        # Filtrando as inscrições por dia se um dia específico for fornecido
        if day:
            inscricoes = inscricoes.filter(inscricaosessao__sessao__dia=day)

        sessoes = Inscricaosessao.objects.filter(inscricao__in=inscricoes)
        if not sessoes.exists():
            return redirect('utilizadores:mensagem6', 1)  # Assumindo que esta URL trata o caso sem sessões


        buffer = io.StringIO()
        writer = csv.writer(buffer, delimiter=';')
        writer.writerow(['Atividade', 'Tipo', 'Data da Sessao', 'Hora de Inicio', 'Hora de Termino', 'Numero de Participantes', 'Numero Maximo de Participantes'])
        populate_csv(writer, inscricoes)

        # Se a opção 'rep' for 'yes', salvar o arquivo internamente
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")  # Gerar timestamp
        if rep == 'yes':
            filepath = os.path.join(RELATORIOS_DIR, f"atividades_dia_aberto_{ano}_{timestamp}.csv")
            with open(filepath, 'w', newline='', encoding='utf-8') as file:
                file.write(buffer.getvalue())

        buffer.seek(0)
        response = HttpResponse(buffer, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="atividades_dia_aberto_{ano}_{timestamp}.csv"'
        return response

    except Diaaberto.DoesNotExist:
        return HttpResponse("Dia Aberto não encontrado para o ano especificado.", status=404)

def populate_csv(writer, inscricoes):
    sessoes = (Inscricaosessao.objects
               .filter(inscricao__in=inscricoes)
               .select_related('sessao', 'sessao__atividadeid', 'sessao__horarioid', 'inscricao')
               .annotate(nparticipantes_total=Sum('nparticipantes'))
               .order_by('inscricao__dia', 'sessao__horarioid__inicio'))

    if not sessoes.exists():
        return redirect('utilizadores:mensagem6', 1)

    sessoes_por_dia = {}
    for inscricaosessao in sessoes:
        dia_inscricao = inscricaosessao.inscricao.dia
        if dia_inscricao not in sessoes_por_dia:
            sessoes_por_dia[dia_inscricao] = []
        sessoes_por_dia[dia_inscricao].append(inscricaosessao)

    for dia, sessoes_dia in sessoes_por_dia.items():
        for inscricaosessao in sessoes_dia:
            atividade = inscricaosessao.sessao.atividadeid
            horario = inscricaosessao.sessao.horarioid
            nparticipantes_total = inscricaosessao.nparticipantes_total
            writer.writerow([
                atividade.nome,
                atividade.tipo,
                dia.strftime('%Y-%m-%d'),
                horario.inicio.strftime('%H:%M'),
                horario.fim.strftime('%H:%M'),
                nparticipantes_total,
                atividade.participantesmaximo
            ])

@handle_db_errors
def relatorio_presencas(request, ano=None, rep=None):
    try:
        dia_aberto = Diaaberto.objects.get(ano=ano)
    except Diaaberto.DoesNotExist:
        return redirect('utilizadores:mensagem2', 20)
    
    inscricoes_presenca = Inscricao.objects.filter(
        diaaberto=dia_aberto
    ).select_related('escola').order_by('dia', 'escola', 'ano', 'turma')

    if not inscricoes_presenca.exists():
        return redirect('utilizadores:mensagem2', 20)
    
    # Grouping inscricoes by day
    inscricoes_grouped_by_day = {}
    for inscricao in inscricoes_presenca:
        day = inscricao.dia
        if day not in inscricoes_grouped_by_day:
            inscricoes_grouped_by_day[day] = []
        inscricoes_grouped_by_day[day].append(inscricao)

    # Creating a context for each day
    context_by_day = []
    for day, inscricoes_list in inscricoes_grouped_by_day.items():
        context_by_day.append({
            'day': day,
            'inscricoes': inscricoes_list,
            'total_nalunos': sum(i.nalunos for i in inscricoes_list),
            'total_presentes': sum(i.presentes for i in inscricoes_list),
        })

    context = {
        'ano': ano,
        'days_context': context_by_day,
    }

    filename = f"presencas_dia_aberto_{ano}.pdf"
    if rep == 'yes':
        # Salvar o PDF internamente e retorná-lo
        return render_pdf(request, "relatorios/relatorio_presencas.html", context, filename, save_file=True)
    else:
        # Apenas gerar o PDF para visualização, sem salvar
        return render_pdf(request, "relatorios/relatorio_presencas.html", context, filename)
@handle_db_errors
def relatorio_presencas_csv(request, ano=None, rep=None):
    # Fetch the 'Diaaberto' instance for the given year
    try:
        dia_aberto = Diaaberto.objects.get(ano=ano)
    except Diaaberto.DoesNotExist:
        return HttpResponse(f"Não existe Dia Aberto para o ano {ano}.", status=404)
    
    # Fetch all 'Inscricao' instances for the 'Diaaberto'
    inscricoes = Inscricao.objects.filter(diaaberto=dia_aberto).select_related('escola')

    if not inscricoes.exists():
        return HttpResponse(f"Não existem inscrições para o Dia Aberto do ano {ano}.", status=404)

    if rep == 'yes':
        create_report(request, dia_aberto, 'CSV', 'Presencas')

    # Prepare the response object with the CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="presencas_dia_aberto_{ano}.csv"'
    writer = csv.writer(response, delimiter=';')

    # Write the header row to the CSV file
    writer.writerow(['Escola', 'Dia','Ano Escolar', 'Turma', 'Hora de Chegada', 'Numero de Inscricoes', 'Numero de Presentes'])

    # Iterate over inscricoes and write their data to the CSV
    for inscricao in inscricoes:
        writer.writerow([
            inscricao.escola.nome,
            inscricao.dia,
            inscricao.ano,
            inscricao.turma,
            inscricao.hora_chegada,
            inscricao.nalunos,
            inscricao.presentes,
        ])

    # Return the response to be downloaded as a CSV file
    return response
@handle_db_errors
def relatorio_view(request):
    if request.method == 'POST':
        # Processar o formulário enviado
        selected_year = request.POST.get('year')
        report_type = request.POST.get('report_type')
        save_to_database = request.POST.get('save_to_database')
        tema = request.POST.get('theme')
        day = request.POST.get('day')
        print(day)


        if tema == 'atividade':
            if day == 'todos':
                if report_type == 'pdf':
                    return HttpResponseRedirect(reverse('relatorios:relatorio_atividade', kwargs={'ano': selected_year, 'rep': save_to_database}))
                elif report_type == 'csv':
                    return HttpResponseRedirect(reverse('relatorios:relatorio_atividade_csv', kwargs={'ano': selected_year, 'rep': save_to_database}))
            else:
                if report_type == 'pdf':
                    return HttpResponseRedirect(reverse('relatorios:relatorio_atividade_dia', kwargs={'ano': selected_year, 'rep': save_to_database, 'day': day}))
                elif report_type == 'csv':
                    return HttpResponseRedirect(reverse('relatorios:relatorio_atividade_csv_dia', kwargs={'ano': selected_year, 'rep': save_to_database, 'day': day}))
        elif tema == 'transporte':
            if day == 'todos':
                if report_type == 'pdf':
                    return HttpResponseRedirect(reverse('relatorios:relatorio_transporte', kwargs={'ano': selected_year, 'rep': save_to_database}))
                elif report_type == 'csv':
                    return HttpResponseRedirect(reverse('relatorios:relatorio_transporte_csv', kwargs={'ano': selected_year, 'rep': save_to_database}))
            else:
                if report_type == 'pdf':
                    return HttpResponseRedirect(reverse('relatorios:relatorio_transporte_dia', kwargs={'ano': selected_year, 'rep': save_to_database, 'day': day}))
                elif report_type == 'csv':
                    return HttpResponseRedirect(reverse('relatorios:relatorio_transporte_csv_dia', kwargs={'ano': selected_year, 'rep': save_to_database, 'day': day}))
        elif tema == 'presencas':
            if report_type == 'pdf':
                return HttpResponseRedirect(reverse('relatorios:relatorio_presencas', kwargs={'ano': selected_year, 'rep': save_to_database}))
            elif report_type == 'csv':
                return HttpResponseRedirect(reverse('relatorios:relatorio_presencas_csv', kwargs={'ano': selected_year, 'rep': save_to_database}))
        elif tema == 'almoco':
            if day == 'todos':
                if report_type == 'pdf':
                    return HttpResponseRedirect(reverse('relatorios:relatorio_almoco_pdf', kwargs={'ano': selected_year, 'rep': save_to_database}))
                elif report_type == 'csv':
                    return HttpResponseRedirect(reverse('relatorios:relatorio_almoco_csv', kwargs={'ano': selected_year, 'rep': save_to_database}))
            elif day == 'Nenhum dia encontrado':
                return redirect('utilizadores:mensagem',8008)
            else:
                if report_type == 'pdf':
                    return HttpResponseRedirect(reverse('relatorios:relatorio_almoco_dia_pdf', kwargs={'ano': selected_year, 'rep': save_to_database, 'day': day}))
                elif report_type == 'csv':
                    return HttpResponseRedirect(reverse('relatorios:relatorio_almoco_dia_csv', kwargs={'ano': selected_year, 'rep': save_to_database, 'day': day}))
        else:
            # Adicione condições para outros temas conforme necessário
            pass

    else:
        # Se o método HTTP não for POST, renderizar o formulário para selecionar o ano e o tipo de relatório
        anos = Diaaberto.objects.values_list('ano', flat=True).order_by('-ano')
        dias_abertos = sorted(set(anos), reverse=True)
        context = {
            'dias_abertos': dias_abertos,  # passar o tema para o template, se necessário
        }
        return render(request, 'relatorios/p_relatorio_atividade.html', context)
    
@handle_db_errors
def abrir_relatorio(request, filename):
    file_path = os.path.join(RELATORIOS_DIR, filename)
    return FileResponse(open(file_path, 'rb'), as_attachment=False)

@handle_db_errors
def deletar_relatorio(request, filename):
    file_path = os.path.join(RELATORIOS_DIR, filename)
    if os.path.exists(file_path):
        os.remove(file_path)
    return HttpResponseRedirect('/relatorios/')

@handle_db_errors
def relatorio_almoco_pdf(request, ano=None, rep=None):
    try:
        dia_aberto = Diaaberto.objects.get(ano=ano)
    except Diaaberto.DoesNotExist:
        return redirect('utilizadores:mensagem',8005)
    
    precoalunos = dia_aberto.precoalunos
    precoprofessores = dia_aberto.precoprofessores
    
    inscricoes = Inscricao.objects.filter(diaaberto_id=dia_aberto.id)
    if not inscricoes.exists():
        return redirect('utilizadores:mensagem',8006)
    
    if not request.user.groups.filter(name="Administrador").exists():
        return redirect('utilizadores:mensagem',8007)

    # Inicialização dos totais acumulados
    total_acumulado = {
        'Penha': {'professores': 0, 'alunos': 0, 'total': 0, 'total_combinado': 0},
        'Gambelas': {'professores': 0, 'alunos': 0, 'total': 0, 'total_combinado': 0},
        'Portimao': {'professores': 0, 'alunos': 0, 'total': 0, 'total_combinado': 0}
    }

    # Dicionário para armazenar os dados por dia
    dados_por_dia = defaultdict(lambda: {
        'inscricoes': [],
        'totais': defaultdict(lambda: {'professores': 0, 'alunos': 0, 'total': 0})
    })

    for inscricao in inscricoes:
        inscricao_prato = Inscricaoprato.objects.filter(inscricao_id=inscricao.id).all()
        for prato in inscricao_prato:
            dia = inscricao.dia
            campus = prato.campus.nome
            dados_por_dia[dia]['inscricoes'].append({
                "inscricaoPrato_id": prato.id,
                "nome_escola": inscricao.escola.nome,
                "local_escola": inscricao.escola.local,
                "dia_inscricao": dia,
                "ano_inscricao": inscricao.ano,
                "turma_inscricao": inscricao.turma,
                "campus": campus,
                "alunos_pratos": prato.npratosalunos,
                "professores_pratos": prato.npratosdocentes,
                "querem_comer": prato.npratosalunos + prato.npratosdocentes
            })

            # Atualização dos totais diários
            dados_por_dia[dia]['totais'][campus]['professores'] += prato.npratosdocentes
            dados_por_dia[dia]['totais'][campus]['alunos'] += prato.npratosalunos
            dados_por_dia[dia]['totais'][campus]['total'] += (prato.npratosalunos + prato.npratosdocentes)

            # Acumulando os totais gerais
            total_acumulado[campus]['professores'] += prato.npratosdocentes
            total_acumulado[campus]['alunos'] += prato.npratosalunos
            total_acumulado[campus]['total'] += (prato.npratosalunos + prato.npratosdocentes)
            total_acumulado[campus]['total_combinado'] += (prato.npratosalunos + prato.npratosdocentes)

    filename = f"almocos_dia_aberto_{ano}.pdf"
    context = {
        'dados_por_dia': dict(dados_por_dia),
        'ano': ano,
        'precoalunos': precoalunos,
        'precoprofessores': precoprofessores,
        'total_acumulado': total_acumulado
    }

    if rep == 'yes':
        return render_pdf(request, "relatorios/relatorio_almoco_pdf.html", context, filename, save_file=True)
    else:
        return render_pdf(request, "relatorios/relatorio_almoco_pdf.html", context, filename, save_file=False)

#normal
@handle_db_errors
def relatorio_almoco_csv2(request, ano=None, rep=None):
    try:
        dia_aberto = Diaaberto.objects.get(ano=ano)
    except Diaaberto.DoesNotExist:
        return redirect('utilizadores:mensagem',8005)

    inscricoes = Inscricao.objects.filter(diaaberto_id=dia_aberto.id).prefetch_related('escola', 'inscricaoprato_set')
    if not inscricoes.exists():
        return redirect('utilizadores:mensagem',8006)

    if not request.user.groups.filter(name="Administrador").exists():
        return redirect('utilizadores:mensagem',8010)

    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=';')
    writer.writerow(['Nº Grupo', 'Escola', 'Ano/Turma', 'Local', 'Dia', 'Cantina Penha', 'Cantina Gambelas', 'Cantina Portimão'])

    inscricoes_por_dia = defaultdict(list)
    for inscricao in inscricoes:
        inscricoes_por_dia[inscricao.dia].append(inscricao)

    for dia, inscricoes_do_dia in inscricoes_por_dia.items():
        totais_por_campus = defaultdict(int)

        for inscricao in inscricoes_do_dia:
            escola = inscricao.escola.nome
            ano_turma = f"{inscricao.ano}/{inscricao.turma if inscricao.turma else ''}"
            local = inscricao.escola.local

            for prato in inscricao.inscricaoprato_set.all():
                pratos_por_campus = defaultdict(int)
                campus = prato.campus.nome
                pratos_por_campus[campus] += prato.npratosalunos + prato.npratosdocentes
                totais_por_campus[campus] += prato.npratosalunos + prato.npratosdocentes

                writer.writerow([
                    prato.id,
                    escola,
                    ano_turma,
                    local,
                    dia,
                    "---" if pratos_por_campus['Penha'] == 0 else pratos_por_campus['Penha'],
                    "---" if pratos_por_campus['Gambelas'] == 0 else pratos_por_campus['Gambelas'],
                    "---" if pratos_por_campus['Portimão'] == 0 else pratos_por_campus['Portimão']
                ])

        # Escrevendo os totais de cada campus para o dia
        writer.writerow([
            'Total', '---', '---', '---', '---',
            totais_por_campus['Penha'],
            totais_por_campus['Gambelas'],
            totais_por_campus['Portimão']
        ])

    # Se a opção 'rep' for 'yes', salvar o arquivo internamente
    if rep == 'yes':
        #create_report(request, dia_aberto, 'CSV', 'Almoco') #salva na base de dados
        filepath = os.path.join(RELATORIOS_DIR, f"almocos_dia_aberto_{ano}.csv")
        with open(filepath, 'w', newline='', encoding='utf-8') as file:
            file.write(buffer.getvalue())
    

    # Retornar o CSV como download
    buffer.seek(0)  # Voltar ao início do buffer
    response = HttpResponse(buffer, content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="almocos_dia_aberto_{ano}.csv"'
    return response

from openpyxl import Workbook
from django.http import HttpResponse
from collections import defaultdict
import os

#novo(excel)
@handle_db_errors
def relatorio_almoco_csv(request, ano=None, rep=None):
    try:
        dia_aberto = Diaaberto.objects.get(ano=ano)
    except Diaaberto.DoesNotExist:
        return redirect('utilizadores:mensagem', 8005)

    inscricoes = Inscricao.objects.filter(diaaberto_id=dia_aberto.id).prefetch_related('escola', 'inscricaoprato_set')
    if not inscricoes.exists():
        return redirect('utilizadores:mensagem', 8006)

    if not request.user.groups.filter(name="Administrador").exists():
        return redirect('utilizadores:mensagem', 8010)

    wb = Workbook()
    wb.remove(wb.active)  # remove a aba padrão criada automaticamente

    sheets_por_dia = {}  # Dicionário para controlar as abas por dia
    totais_gerais = defaultdict(int)
    totais_por_campus_dia = defaultdict(lambda: defaultdict(int))  # Dicionário para controlar totais por dia e campus

    # Criar a aba "Total" primeiro para reorganizar depois
    total_sheet = wb.create_sheet(title="Total")
    total_sheet.append(['Nº Grupo', 'Escola', 'Ano/Turma', 'Local', 'Dia', 'Cantina Penha', 'Cantina Gambelas', 'Cantina Portimão'])

    for inscricao in inscricoes:
        dia = inscricao.dia
        if dia not in sheets_por_dia:
            ws = wb.create_sheet(title=f"{dia}")
            ws.append(['Nº Grupo', 'Escola', 'Ano/Turma', 'Local', 'Dia', 'Cantina Penha', 'Cantina Gambelas', 'Cantina Portimão'])
            sheets_por_dia[dia] = ws

        ws = sheets_por_dia[dia]
        escola = inscricao.escola.nome
        ano_turma = f"{inscricao.ano}/{inscricao.turma if inscricao.turma else ''}"
        local = inscricao.escola.local

        for prato in inscricao.inscricaoprato_set.all():
            campus = prato.campus.nome
            quantidade_pratos = prato.npratosalunos + prato.npratosdocentes
            totais_por_campus_dia[dia][campus] += quantidade_pratos
            totais_gerais[campus] += quantidade_pratos

            line = [prato.id, escola, ano_turma, local, dia,
                    "---" if campus == 'Penha' and quantidade_pratos == 0 else quantidade_pratos if campus == 'Penha' else "---",
                    "---" if campus == 'Gambelas' and quantidade_pratos == 0 else quantidade_pratos if campus == 'Gambelas' else "---",
                    "---" if campus == 'Portimão' and quantidade_pratos == 0 else quantidade_pratos if campus == 'Portimão' else "---"]

            ws.append(line)
            total_sheet.append(line)

    # Adicionar os totais de cada campus no final de cada worksheet
    for dia, ws in sheets_por_dia.items():
        ws.append([
            'Total', '---', '---', '---', '---',
            totais_por_campus_dia[dia]['Penha'],
            totais_por_campus_dia[dia]['Gambelas'],
            totais_por_campus_dia[dia]['Portimão']
        ])

    # Adicionar os totais gerais na aba 'Total'
    total_sheet.append([
        'Total Geral', '---', '---', '---', '---',
        totais_gerais['Penha'],
        totais_gerais['Gambelas'],
        totais_gerais['Portimão']
    ])

    wb._sheets.append(wb._sheets.pop(wb._sheets.index(total_sheet)))

    if rep == 'yes':
        file_path = os.path.join(RELATORIOS_DIR, f"almocos_dia_aberto_{ano}.xlsx")
        wb.save(file_path)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="almocos_dia_aberto_{ano}.xlsx"'
    wb.save(response)

    return response



@handle_db_errors
def inscricoes_por_dia(request, tema, ano):
    if tema == "almoco":
        dia_aberto = Diaaberto.objects.filter(ano=ano).first()
        if dia_aberto:
            inscricoes = Inscricao.objects.filter(
            diaaberto_id=dia_aberto.id,
            inscricaoprato__isnull=False  # Este filtro garante que só consideremos as inscrições que têm pratos
            ).values_list('dia', flat=True).distinct()
            return JsonResponse({'dias': list(inscricoes)}, safe=False)
    if tema == 'atividade':
        dia_aberto = Diaaberto.objects.filter(ano=ano).first()
        if dia_aberto:
            inscricoes = Inscricao.objects.filter(diaaberto_id=dia_aberto.id).values_list('dia', flat=True).distinct()
            return JsonResponse({'dias': list(inscricoes)}, safe=False)
    if tema == 'transporte':
        dia_aberto = Diaaberto.objects.filter(ano=ano).first()
        if dia_aberto:
            inscricoes = Inscricao.objects.filter(diaaberto_id=dia_aberto.id).values_list('dia', flat=True).distinct()
            return JsonResponse({'dias': list(inscricoes)}, safe=False)
    return JsonResponse({'error': 'Nenhum dado encontrado'}, status=404)

@handle_db_errors
def relatorio_almoco_dia_pdf(request, ano=None, rep=None, day=None):
    try:
        dia_aberto = Diaaberto.objects.get(ano=ano)
    except Diaaberto.DoesNotExist:
        return redirect('utilizadores:mensagem',8005)

    if not request.user.groups.filter(name="Administrador").exists():
        return redirect('utilizadores:mensagem',8007)
    
    if not day:
        return redirect('utilizadores:mensagem',8009)

    inscricoes = Inscricao.objects.filter(diaaberto_id=dia_aberto.id, dia=day)
    if not inscricoes.exists():
        return redirect('utilizadores:mensagem',8006)

    dados_por_dia = defaultdict(lambda: {'inscricoes': [], 'totais': defaultdict(lambda: {'professores': 0, 'alunos': 0, 'total': 0})})

    for inscricao in inscricoes:
        for prato in inscricao.inscricaoprato_set.all():
            dados_por_dia[day]['inscricoes'].append({
                "inscricaoPrato_id": prato.id,
                "nome_escola": inscricao.escola.nome,
                "local_escola": inscricao.escola.local,
                "dia_inscricao": inscricao.dia,
                "ano_inscricao": inscricao.ano,
                "turma_inscricao": inscricao.turma,
                "campus": prato.campus.nome,
                "alunos_pratos": prato.npratosalunos,
                "professores_pratos": prato.npratosdocentes,
                "querem_comer": prato.npratosalunos + prato.npratosdocentes
            })
            campus_totals = dados_por_dia[day]['totais'].get(prato.campus.nome, {'professores': 0, 'alunos': 0, 'total': 0})
            campus_totals['professores'] += prato.npratosdocentes
            campus_totals['alunos'] += prato.npratosalunos
            campus_totals['total'] += prato.npratosalunos + prato.npratosdocentes
            dados_por_dia[day]['totais'][prato.campus.nome] = campus_totals

    context = {
        'dados_por_dia': dict(dados_por_dia),  # Convertendo defaultdict para dict
        'ano': ano,
        'precoalunos': dia_aberto.precoalunos,
        'precoprofessores': dia_aberto.precoprofessores,
        'selected_day': day  # Informação sobre o dia selecionado
    }

    filename = f"almocos_dia_aberto_{ano}_{day}.pdf"
    if rep == 'yes':
        return render_pdf(request, "relatorios/relatorio_almoco_dia_pdf.html", context, filename, save_file=True)
    else:
        return render_pdf(request, "relatorios/relatorio_almoco_dia_pdf.html", context, filename, save_file=False)
    
#normal
@handle_db_errors
def relatorio_almoco_dia_csv2(request, ano=None, rep=None, day=None):
    try:
        dia_aberto = Diaaberto.objects.get(ano=ano)
    except Diaaberto.DoesNotExist:
        return redirect('utilizadores:mensagem',8005)
    print("Ano: ", ano ,"Dia: ", day)  # Adicione isso para depuração

    if not request.user.groups.filter(name="Administrador").exists():
        return redirect('utilizadores:mensagem',8010)

    # Filtra as inscrições por dia específico
    inscricoes = Inscricao.objects.filter(diaaberto_id=dia_aberto.id, dia=day).prefetch_related('escola', 'inscricaoprato_set')
    if not inscricoes.exists():
        return redirect('utilizadores:mensagem',8006)

    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=';')
    writer.writerow(['Nº Grupo', 'Escola', 'Ano/Turma', 'Local', 'Dia', 'Cantina Penha', 'Cantina Gambelas', 'Cantina Portimão'])

    totais_por_campus = defaultdict(int)

    for inscricao in inscricoes:
        escola = inscricao.escola.nome
        ano_turma = f"{inscricao.ano}/{inscricao.turma if inscricao.turma else ''}"
        local = inscricao.escola.local

        for prato in inscricao.inscricaoprato_set.all():
            pratos_por_campus = defaultdict(int)
            campus = prato.campus.nome
            pratos_por_campus[campus] += prato.npratosalunos + prato.npratosdocentes
            totais_por_campus[campus] += prato.npratosalunos + prato.npratosdocentes

            writer.writerow([
                prato.id,
                escola,
                ano_turma,
                local,
                day,
                "---" if pratos_por_campus['Penha'] == 0 else pratos_por_campus['Penha'],
                "---" if pratos_por_campus['Gambelas'] == 0 else pratos_por_campus['Gambelas'],
                "---" if pratos_por_campus['Portimão'] == 0 else pratos_por_campus['Portimão']
            ])

    # Escrevendo os totais de cada campus para o dia
    writer.writerow([
        'Total', '---', '---', '---', '---',
        totais_por_campus['Penha'],
        totais_por_campus['Gambelas'],
        totais_por_campus['Portimão']
    ])

    # Se a opção 'rep' for 'yes', salvar o arquivo internamente
    if rep == 'yes':
        filepath = os.path.join(RELATORIOS_DIR, f"almocos_dia_aberto_{ano}_{day}.csv")
        with open(filepath, 'w', newline='', encoding='utf-8') as file:
            file.write(buffer.getvalue())

    # Retornar o CSV como download
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="almocos_dia_aberto_{ano}_{day}.csv"'
    return response


#novo(excel)
@handle_db_errors
def relatorio_almoco_dia_csv(request, ano=None, rep=None, day=None):
    try:
        dia_aberto = Diaaberto.objects.get(ano=ano)
    except Diaaberto.DoesNotExist:
        return redirect('utilizadores:mensagem', 8005)

    if not request.user.groups.filter(name="Administrador").exists():
        return redirect('utilizadores:mensagem', 8010)

    # Filtra as inscrições por dia específico
    inscricoes = Inscricao.objects.filter(diaaberto_id=dia_aberto.id, dia=day).prefetch_related('escola', 'inscricaoprato_set')
    if not inscricoes.exists():
        return redirect('utilizadores:mensagem', 8006)

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Almoços"

    ws.append(['Nº Grupo', 'Escola', 'Ano/Turma', 'Local', 'Dia', 'Cantina Penha', 'Cantina Gambelas', 'Cantina Portimão'])
    totais_por_campus = defaultdict(int)

    for inscricao in inscricoes:
        escola = inscricao.escola.nome
        ano_turma = f"{inscricao.ano}/{inscricao.turma if inscricao.turma else ''}"
        local = inscricao.escola.local

        for prato in inscricao.inscricaoprato_set.all():
            pratos_por_campus = defaultdict(int)
            campus = prato.campus.nome
            pratos_por_campus[campus] = prato.npratosalunos + prato.npratosdocentes
            totais_por_campus[campus] += pratos_por_campus[campus]

            ws.append([
                prato.id,
                escola,
                ano_turma,
                local,
                day,
                "---" if pratos_por_campus['Penha'] == 0 else pratos_por_campus['Penha'],
                "---" if pratos_por_campus['Gambelas'] == 0 else pratos_por_campus['Gambelas'],
                "---" if pratos_por_campus['Portimão'] == 0 else pratos_por_campus['Portimão']
            ])

    # Escrevendo os totais de cada campus para o dia
    ws.append([
        'Total', '---', '---', '---', '---',
        totais_por_campus['Penha'],
        totais_por_campus['Gambelas'],
        totais_por_campus['Portimão']
    ])

    # Se a opção 'rep' for 'yes', salvar o arquivo internamente
    if rep == 'yes':
        filepath = os.path.join(RELATORIOS_DIR, f"almocos_dia_aberto_{ano}_{day}.xlsx")
        wb.save(filepath)

    # Preparar o arquivo para download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="almocos_dia_aberto_{ano}_{day}.xlsx"'
    wb.save(response)

    return response


@handle_db_errors
def get_anos_disponiveis_almocos(request):
    # Buscar anos onde existem inscrições prato associadas ao tema "almoços"
    anos = Diaaberto.objects.filter(
        inscricao__inscricaoprato__isnull=False
    ).distinct().order_by('ano').values_list('ano', flat=True)

    return JsonResponse(list(anos), safe=False)
@handle_db_errors
def get_all_years(request):
    # Buscar anos onde existem inscrições prato associadas ao tema "almoços"
    anos = Diaaberto.objects.all().order_by('ano').values_list('ano', flat=True).distinct()

    return JsonResponse(list(anos), safe=False)







